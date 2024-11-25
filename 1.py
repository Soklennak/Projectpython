import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import time
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import threading
from openpyxl import load_workbook  # Import to handle Excel files

# Initialize the main window
root = tk.Tk()
root.title("Email Scheduler")
root.geometry("500x400")
root.configure(bg="#f0f8ff")  # Light blue background

# Global variables for the Excel file and email credentials
email_user = ""  # This will be set by the user in the interface
email_password = ""  # Password entered by the user at runtime
excel_file_path = ""  # Path to the Excel file

# SMTP Server Configuration
smtp_server = "smtp.gmail.com"
smtp_port = 587


# Function to select the Excel file
def browse_file():
    global excel_file_path
    excel_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    excel_file_label.config(text=excel_file_path)


# Function to send email
def send_email(to_email, subject, body):
    try:
        # Create Email Message
        message = MIMEMultipart()
        message["From"] = email_user
        message["To"] = to_email
        message["Subject"] = subject
        message.attach(MIMEText(body, "plain"))

        # Connect to SMTP Server and Send Email
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(email_user, email_password)  # User-provided password
            server.send_message(message)
            print(f"Email sent to {to_email}")
    except Exception as e:
        print(f"Failed to send email to {to_email}: {e}")


# Function to schedule email
def schedule_email(scheduled_time, subject, body):
    while True:
        # Check if current time matches the specified time
        now = datetime.now()
        if now >= scheduled_time:
            # Load Excel File and Read Data
            try:
                workbook = load_workbook(excel_file_path)
                sheet = workbook.active

                # Read only name and email columns (Assuming name is in column 1 and email in column 2)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    name, email = row  # Read only name and email columns
                    if name and email:
                        personalized_body = body.replace("{name}", name)  # Optional personalization
                        send_email(email, subject, personalized_body)

                messagebox.showinfo("Success", "Emails sent successfully!")
                break  # Exit the loop after sending emails
            except FileNotFoundError:
                messagebox.showerror("Error", "Excel file not found.")
                break
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read Excel file or send emails: {e}")
                break
        time.sleep(1)  # Avoid excessive CPU usage


# Function to start the scheduling process
def start_scheduling():
    global email_user, email_password

    # Get sender's email
    email_user = sender_email_entry.get()

    if not email_user:
        messagebox.showerror("Error", "Please enter sender's email.")
        return

    if not excel_file_path:
        messagebox.showerror("Error", "Please select an Excel file.")
        return

    # Get the password
    email_password = password_entry.get()

    if not email_password:
        messagebox.showerror("Error", "Please enter your email password.")
        return

    try:
        # Get the scheduling date and time
        scheduled_date = datetime.strptime(schedule_date_entry.get(), "%Y-%m-%d")
        scheduled_time = datetime.strptime(schedule_time_entry.get(), "%H:%M")
        scheduled_time = scheduled_date.replace(hour=scheduled_time.hour, minute=scheduled_time.minute)

        if scheduled_time <= datetime.now():
            messagebox.showerror("Error", "Scheduled time must be in the future.")
            return

        subject = subject_entry.get()
        body = body_entry.get("1.0", "end-1c")

        # Start the email scheduler in a separate thread
        scheduling_thread = threading.Thread(target=schedule_email, args=(scheduled_time, subject, body))
        scheduling_thread.start()

        print(f"Email will be sent at {scheduled_time}")

    except ValueError:
        messagebox.showerror("Error", "Invalid date/time format. Please use YYYY-MM-DD for date and HH:MM for time.")


# Create the interface elements with colors
tk.Label(root, text="Sender Email:", bg="#f0f8ff", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
sender_email_entry = tk.Entry(root, width=40, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
sender_email_entry.pack(pady=5)

tk.Label(root, text="Email Password:", bg="#f0f8ff", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
password_entry = tk.Entry(root, width=40, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50", show="*")
password_entry.pack(pady=5)

# Excel file selection
tk.Label(root, text="Select Excel File:", bg="#f0f8ff", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
excel_file_label = tk.Label(root, text="No file selected", bg="#f0f8ff", font=("Helvetica", 10), fg="#2c3e50")
excel_file_label.pack(pady=5)
browse_button = tk.Button(root, text="Browse", command=browse_file, font=("Helvetica", 10), bg="#3498db", fg="white")
browse_button.pack(pady=5)

# Subject input
tk.Label(root, text="Subject Line:", bg="#f0f8ff", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
subject_entry = tk.Entry(root, width=40, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
subject_entry.pack(pady=5)

# Email body input
tk.Label(root, text="Email Body:", bg="#f0f8ff", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
body_entry = tk.Text(root, width=40, height=5, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
body_entry.pack(pady=5)

# Schedule date and time input
tk.Label(root, text="Schedule Date (YYYY-MM-DD):", bg="#f0f8ff", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
schedule_date_entry = tk.Entry(root, width=40, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
schedule_date_entry.pack(pady=5)

tk.Label(root, text="Schedule Time (HH:MM):", bg="#f0f8ff", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
schedule_time_entry = tk.Entry(root, width=40, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
schedule_time_entry.pack(pady=5)

# Start Scheduler Button
start_button = tk.Button(root, text="Start Scheduler", command=start_scheduling, bg="#27ae60", fg="white", font=("Helvetica", 12))
start_button.pack(pady=20)

root.mainloop()
