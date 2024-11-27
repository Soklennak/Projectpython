import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from openpyxl import load_workbook
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import time
import os

# Initialize the main window
root = tk.Tk()
root.title("Email Scheduler")
root.geometry("600x700")  # Adjusted window height
root.configure(bg="#f7f9fc")  # Light background for a clean UI
root.resizable(width="false", height="false")

# Global variables for the Excel file, email credentials, and attachments
email_user = ""
email_password = ""
excel_file_path = ""
attachments = []

# SMTP Server Configuration
smtp_server = "smtp.gmail.com"
smtp_port = 587

# Function to select the Excel file
def browse_file():
    global excel_file_path
    excel_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    excel_file_label.config(text=excel_file_path if excel_file_path else "No file selected")

# Function to select attachments
def browse_attachments():
    global attachments
    files = filedialog.askopenfilenames(filetypes=[("All Files", "*.*")])
    attachments.extend(files)
    attachment_label.config(text="Attachments: " + ", ".join(os.path.basename(file) for file in attachments))

# Function to send email with attachments
def send_email(to_email, subject, body):
    try:
        message = MIMEMultipart()
        message["From"] = email_user
        message["To"] = to_email
        message["Subject"] = subject
        message.attach(MIMEText(body, "plain"))

        # Attach all files
        for file_path in attachments:
            with open(file_path, "rb") as file:
                part = MIMEApplication(file.read(), Name=os.path.basename(file_path))
                part['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
                message.attach(part)

        # Send email via SMTP server
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(email_user, email_password)
            server.send_message(message)
            print(f"Email sent to {to_email}")
    except Exception as e:
        print(f"Failed to send email to {to_email}: {e}")

# Function to schedule and send multiple emails at different times
def schedule_email(scheduled_time):
    while True:
        now = datetime.now()
        if now >= scheduled_time:
            try:
                workbook = load_workbook(excel_file_path)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 4 and row[1]:
                        name = row[0] or "User"
                        email = row[1]
                        subject = row[2]
                        body = row[3]
                        
                        # Personalize email with recipient's name
                        personalized_body = body.replace("{name}", name)
                        
                        send_email(email, subject, personalized_body)
                        print(f"Sent email to {email} with subject: {subject}")

                messagebox.showinfo("Success", "Emails sent successfully!")
                return  # Exit after sending all emails for the current scheduled time
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read Excel file or send emails: {e}")
                return
        time.sleep(30)  # Sleep for a short time before checking the time again

# Function to start the scheduling process
def start_scheduling():
    global email_user, email_password, attachments

    email_user = sender_email_entry.get()
    if not email_user:
        messagebox.showerror("Error", "Please enter sender's email.")
        return

    if not excel_file_path:
        messagebox.showerror("Error", "Please select an Excel file.")
        return

    email_password = password_entry.get()
    if not email_password:
        messagebox.showerror("Error", "Please enter your email password.")
        return

    try:
        scheduled_date = datetime.strptime(schedule_date_entry.get(), "%Y-%m-%d")  # Single date
        scheduled_time = datetime.strptime(schedule_time_entry.get(), "%H:%M")
        scheduled_time = scheduled_date.replace(hour=scheduled_time.hour, minute=scheduled_time.minute)

        if scheduled_time <= datetime.now():
            messagebox.showerror("Error", "Scheduled time must be in the future.")
            return

        # Starting a separate thread to handle email sending
        scheduling_thread = threading.Thread(target=schedule_email, args=(scheduled_time,))
        scheduling_thread.start()

        messagebox.showinfo("Scheduled", f"Emails will be sent at {scheduled_time}.")

    except ValueError:
        messagebox.showerror("Error", "Invalid date/time format. Use YYYY-MM-DD for date and HH:MM for time.")

# UI Elements with better spacing and design
tk.Label(root, text="Email Scheduler", bg="#f7f9fc", fg="#34495e", font=("Helvetica", 16, "bold")).pack(pady=10)

tk.Label(root, text="Sender Email:", bg="#f7f9fc", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
sender_email_entry = tk.Entry(root, width=45, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
sender_email_entry.pack(pady=5)

tk.Label(root, text="Email Password:", bg="#f7f9fc", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
password_entry = tk.Entry(root, width=45, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50", show="*")
password_entry.pack(pady=5)

tk.Label(root, text="Select Excel File:", bg="#f7f9fc", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
excel_file_label = tk.Label(root, text="No file selected", bg="#f7f9fc", font=("Helvetica", 10), fg="#2c3e50")
excel_file_label.pack(pady=5)
browse_button = tk.Button(root, text="Browse", command=browse_file, font=("Helvetica", 10), bg="#3498db", fg="white")
browse_button.pack(pady=5)

tk.Label(root, text="Select Attachments:", bg="#f7f9fc", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
attachment_label = tk.Label(root, text="No files selected", bg="#f7f9fc", font=("Helvetica", 10), fg="#2c3e50")
attachment_label.pack(pady=5)
attachment_button = tk.Button(root, text="Browse", command=browse_attachments, font=("Helvetica", 10), bg="#3498db", fg="white")
attachment_button.pack(pady=5)

tk.Label(root, text="Schedule Date (YYYY-MM-DD):", bg="#f7f9fc", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
schedule_date_entry = tk.Entry(root, width=45, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
schedule_date_entry.pack(pady=5)

tk.Label(root, text="Schedule Time (HH:MM):", bg="#f7f9fc", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
schedule_time_entry = tk.Entry(root, width=45, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
schedule_time_entry.pack(pady=5)

start_button = tk.Button(root, text="Start Scheduler", command=start_scheduling, bg="#27ae60", fg="white", font=("Helvetica", 12, "bold"))
start_button.pack(pady=20)

root.mainloop()
