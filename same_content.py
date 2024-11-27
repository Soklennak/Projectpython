import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import time

# Initialize the main window
root = tk.Tk()
root.title("Email Scheduler")
root.geometry("600x700")
root.configure(bg="#a9cce3")  # Light background for a clean UI
root.resizable(width="false", height="false")

# Set the application icon (replace with the correct icon file path)
try:
    icon = tk.PhotoImage(file="image.png")  # Ensure 'image.png' exists in the same directory
    root.iconphoto(True, icon)
except Exception as e:
    print(f"Failed to load icon: {e}")

# Global variables for the Excel file, attachment, and email credentials
email_user = ""
email_password = ""
excel_file_path = ""
attachment_file_path = ""

# SMTP Server Configuration
smtp_server = "smtp.gmail.com"
smtp_port = 587

# Function to select the Excel file
def browse_file():
    global excel_file_path
    excel_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    excel_file_label.config(text=excel_file_path if excel_file_path else "No file selected")

# Function to select an attachment file
def browse_attachment():
    global attachment_file_path
    attachment_file_path = filedialog.askopenfilename(filetypes=[
        ("All Files", "*.*"),
        ("PDF Files", "*.pdf"),
        ("Image Files", "*.jpg;*.png;*.jpeg"),
        ("Word Documents", "*.doc;*.docx"),
        ("Excel Files", "*.xls;*.xlsx"),
        ("PowerPoint Files", "*.ppt;*.pptx"),])
    attachment_label.config(text=attachment_file_path if attachment_file_path else "No file selected")

# Function to send email
def send_email(to_email, subject, body):
    try:
        message = MIMEMultipart()
        message["From"] = email_user
        message["To"] = to_email
        message["Subject"] = subject
        message.attach(MIMEText(body, "plain"))

        # Attach the file if a valid path is provided
        if attachment_file_path:
            try:
                with open(attachment_file_path, "rb") as attachment:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename={attachment_file_path.split('/')[-1]}"
                )
                message.attach(part)
            except Exception as e:
                print(f"Failed to attach file: {e}")

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(email_user, email_password)
            server.send_message(message)
            print(f"Email sent to {to_email}")
    except Exception as e:
        print(f"Failed to send email to {to_email}: {e}")

# Function to schedule email
def schedule_email(scheduled_time, subject, body):
    while True:
        now = datetime.now()
        if now >= scheduled_time:
            try:
                workbook = load_workbook(excel_file_path)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 2 and row[1]:
                        name = row[0] or "User"
                        email = row[1]
                        personalized_body = body.replace("{name}", name)
                        send_email(email, subject, personalized_body)
                    else:
                        print(f"Skipping invalid row: {row}")

                messagebox.showinfo("Success", "Emails sent successfully!")
                break
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read Excel file or send emails: {e}")
                break
        time.sleep(1)

# Function to start the scheduling process
def start_scheduling():
    global email_user, email_password

    email_user = sender_email_entry.get()
    if not email_user:
        messagebox.showerror("Error", "Please enter sender's email.")
        return

    if not excel_file_path:
        messagebox.showerror("Error", "Please select an Excel file.")
        return

    email_password = password_entry.get()
    if not email_password:
        messagebox.showerror("Error", "Please enter your email password.")
        return

    try:
        scheduled_date = datetime.strptime(schedule_date_entry.get(), "%Y-%m-%d")
        scheduled_time = datetime.strptime(schedule_time_entry.get(), "%H:%M")
        scheduled_time = scheduled_date.replace(hour=scheduled_time.hour, minute=scheduled_time.minute)

        if scheduled_time <= datetime.now():
            messagebox.showerror("Error", "Scheduled time must be in the future.")
            return

        subject = subject_entry.get()
        body = body_entry.get("1.0", "end-1c")

        scheduling_thread = threading.Thread(target=schedule_email, args=(scheduled_time, subject, body))
        scheduling_thread.start()

        messagebox.showinfo("Scheduled", f"Emails will be sent at {scheduled_time}.")

    except ValueError:
        messagebox.showerror("Error", "Invalid date/time format. Use YYYY-MM-DD for date and HH:MM for time.")

# UI Elements
tk.Label(root, text="Email Scheduler", bg="#a9cce3", fg="#34495e", font=("Helvetica", 16, "bold")).pack(pady=10)

tk.Label(root, text="Sender Email:", bg="#a9cce3", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
sender_email_entry = tk.Entry(root, width=45, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
sender_email_entry.pack(pady=5)

tk.Label(root, text="Email Password:", bg="#a9cce3", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
password_entry = tk.Entry(root, width=45, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50", show="*")
password_entry.pack(pady=5)

# Frame for Excel file and attachment selection
file_frame = tk.Frame(root, bg="#a9cce3")
file_frame.pack(pady=10)

tk.Label(file_frame, text="Select Excel File:", bg="#a9cce3", font=("Helvetica", 12), fg="#2c3e50").grid(row=0, column=0, padx=5, pady=5)
excel_file_label = tk.Label(file_frame, text="No file selected", bg="#a9cce3", font=("Helvetica", 10), fg="#2c3e50")
excel_file_label.grid(row=0, column=1, padx=5, pady=5)
browse_button = tk.Button(file_frame, text="Browse", command=browse_file, font=("Helvetica", 10), bg="green", fg="white")
browse_button.grid(row=0, column=2, padx=5, pady=5)

tk.Label(file_frame, text="Attachment:", bg="#a9cce3", font=("Helvetica", 12), fg="#2c3e50").grid(row=1, column=0, padx=5, pady=5)
attachment_label = tk.Label(file_frame, text="No file selected", bg="#a9cce3", font=("Helvetica", 10), fg="#2c3e50")
attachment_label.grid(row=1, column=1, padx=5, pady=5)
attachment_button = tk.Button(file_frame, text="Browse", command=browse_attachment, font=("Helvetica", 10), bg="green", fg="white")
attachment_button.grid(row=1, column=2, padx=5, pady=5)

tk.Label(root, text="Subject Line:", bg="#a9cce3", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
subject_entry = tk.Entry(root, width=45, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
subject_entry.pack(pady=5)

tk.Label(root, text="Email Body:", bg="#a9cce3", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
body_entry = tk.Text(root, width=45, height=5, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
body_entry.pack(pady=5)

tk.Label(root, text="Schedule Date (YYYY-MM-DD):", bg="#a9cce3", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
schedule_date_entry = tk.Entry(root, width=45, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
schedule_date_entry.pack(pady=5)

tk.Label(root, text="Schedule Time (HH:MM):", bg="#a9cce3", font=("Helvetica", 12), fg="#2c3e50").pack(pady=5)
schedule_time_entry = tk.Entry(root, width=45, font=("Helvetica", 10), bg="#ecf0f1", fg="#2c3e50")
schedule_time_entry.pack(pady=5)

# Frame for action buttons
action_frame = tk.Frame(root, bg="#a9cce3")
action_frame.pack(pady=20)

start_button = tk.Button(action_frame, text="Start Scheduler", command=start_scheduling, bg="blue", fg="white", font=("Helvetica", 12, "bold"))
start_button.grid(row=0, column=0, padx=10)

exit_button = tk.Button(action_frame, text="Exit", command=root.destroy, bg="red", fg="white", font=("Helvetica", 12, "bold"))
exit_button.grid(row=0, column=1, padx=10)

root.mainloop()
