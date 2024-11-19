import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook

import time
from datetime import datetime




# Load Excel File
excel_file = "Book2.xlsx"  # Replace with your Excel file path
workbook = load_workbook(excel_file)
sheet = workbook.active

# SMTP Server Configuration
smtp_server = "smtp.gmail.com"  # Replace with your SMTP server
smtp_port = 587
email_user = "sreyvang.phon@student.passerellesnumeriques.org"  # Replace with your email
email_password = "whwn crbx rvgn zqui"  # Replace with your email password or app password


    


def schedule_email(date_time):
    while True:
        # Check if current time matches the specified time
        now = datetime.now()
        if now >= date_time:
            
            # Email Sending Function
            def send_email(to_email, subject, body):
                try:
                    # Create Email Message
                    message = MIMEMultipart()
                    message["From"] = email_user
                    message["To"] = to_email
                    message["Subject"] = subject
                    message.attach(MIMEText(body, "plain"))

                    # Connect to SMTP Server and Send Email
                    with smtplib.SMTP(smtp_server, smtp_port) as server:
                        server.starttls()
                        server.login(email_user, email_password)
                        server.send_message(message)
                        print(f"Email sent to {to_email}")
                except Exception as e:
                    print(f"Failed to send email to {to_email}: {e}")

            # Read Data from Excel and Send Emails
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                name, email, subject, body = row
                personalized_body = body.replace("{name}", name)  # Optional personalization
                send_email(email, subject, personalized_body)
            
            break  # Exit the loop after printing
        time.sleep(1)  # Avoid excessive CPU usage

# Set the date and time for sending the email
# Example: November 19, 2024, at 14:9
scheduled_time = datetime(2024, 11, 19, 16, 10)

# Start the scheduler
print(f"Email will be sent at {scheduled_time}")
schedule_email(scheduled_time)






